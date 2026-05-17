"""Tests for the FastAPI routes — uses httpx TestClient.

These tests run with auth disabled (no Azure AD credentials in env)
and Cosmos disabled (local JSON fallback).
"""

import json
import os
import pytest
from unittest.mock import patch, MagicMock
from fastapi.testclient import TestClient

# Ensure auth and cosmos are disabled for tests
os.environ.setdefault("AZURE_OPENAI_ENDPOINT", "https://test.openai.azure.com/")
os.environ.setdefault("AZURE_OPENAI_API_KEY", "test-key")
os.environ.setdefault("AZURE_TENANT_ID", "")
os.environ.setdefault("AZURE_CLIENT_ID", "")
os.environ.setdefault("COSMOS_ENDPOINT", "")
os.environ.setdefault("COSMOS_KEY", "")

from backend.main import app


client = TestClient(app)


class TestHealth:
    def test_health_returns_ok(self):
        response = client.get("/health")
        assert response.status_code == 200
        assert response.json() == {"status": "ok"}


class TestSchemas:
    def test_list_schemas_empty(self):
        response = client.get("/schemas")
        assert response.status_code == 200
        data = response.json()
        assert "schemas" in data

    def test_create_and_list_schema(self):
        schema = {
            "name": "Test Schema",
            "fields": [
                {
                    "name": "company_name",
                    "field_type": "string",
                    "description": "The company name",
                    "required": True,
                }
            ],
        }
        response = client.post("/schemas", json=schema)
        assert response.status_code == 200
        created = response.json()
        assert created["name"] == "Test Schema"
        assert "id" in created
        assert created["version"] == 1

        # List should include it
        list_response = client.get("/schemas")
        schemas = list_response.json()["schemas"]
        assert any(s["id"] == created["id"] for s in schemas)

    def test_update_schema_increments_version(self):
        schema = {
            "name": "Versioned Schema",
            "fields": [
                {
                    "name": "col1",
                    "field_type": "string",
                    "required": True,
                }
            ],
        }
        response = client.post("/schemas", json=schema)
        created = response.json()
        assert created["version"] == 1

        # Update
        updated_schema = {
            "name": "Versioned Schema Updated",
            "fields": [
                {
                    "name": "col1",
                    "field_type": "string",
                    "required": True,
                },
                {
                    "name": "col2",
                    "field_type": "number",
                    "required": False,
                },
            ],
        }
        update_response = client.put(f"/schemas/{created['id']}", json=updated_schema)
        assert update_response.status_code == 200
        updated = update_response.json()
        assert updated["version"] == 2

    def test_schema_history(self):
        schema = {
            "name": "History Schema",
            "fields": [
                {
                    "name": "col1",
                    "field_type": "string",
                    "required": True,
                }
            ],
        }
        response = client.post("/schemas", json=schema)
        created = response.json()

        # Update it
        schema["name"] = "History Schema v2"
        client.put(f"/schemas/{created['id']}", json=schema)

        # Check history
        history_response = client.get(f"/schemas/{created['id']}/history")
        assert history_response.status_code == 200
        history = history_response.json()
        assert len(history["versions"]) == 2

    def test_delete_nonexistent_schema(self):
        response = client.delete("/schemas/nonexistent_id")
        assert response.status_code == 404


class TestIngest:
    """Test the /ingest endpoint with a mocked LLM call."""

    @patch("backend.main.infer_mapping")
    @patch("backend.main.validate_extraction")
    def test_ingest_basic(self, mock_validate, mock_infer):
        from backend.models import (
            ExcelMapping,
            ColumnMapping,
            Transform,
            ValidationResult,
        )

        # Mock the LLM responses
        mock_infer.return_value = ExcelMapping(
            sheet_name="Sheet1",
            header_row=1,
            data_start_row=2,
            mappings=[
                ColumnMapping(
                    source_col="A",
                    target_field="company_name",
                    transform=Transform.STRIP,
                    notes="Column A is 'Company'",
                ),
            ],
            reasoning="Found headers at row 1",
        )
        mock_validate.return_value = ValidationResult(
            confidence=0.95,
            passed=True,
            summary="All good",
        )

        # Create a minimal xlsx file
        import openpyxl
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Company"
        ws["A2"] = "  Tesco PLC  "
        ws["A3"] = "  Sainsbury's  "

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        schema = {
            "name": "Test",
            "fields": [
                {
                    "name": "company_name",
                    "field_type": "string",
                    "description": "Company name",
                    "required": True,
                }
            ],
        }

        response = client.post(
            "/ingest",
            files={
                "file": (
                    "test.xlsx",
                    buf,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            params={
                "schema_name": "Test",
                "schema_json": json.dumps(schema),
            },
        )

        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True
        assert data["row_count"] == 2
        assert len(data["data"]) == 2
        assert data["data"][0]["company_name"] == "Tesco PLC"
        assert data["data"][1]["company_name"] == "Sainsbury's"
        # Verify per-sheet structure
        assert "sheets" in data
        assert "sheet_names" in data
        assert len(data["sheet_names"]) >= 1
        assert data["schema_version"] >= 1
